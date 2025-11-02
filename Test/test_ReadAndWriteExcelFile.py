import openpyxl

from Utils.BaseClass import BaseClass

class TestReadAndWriteExcelFile(BaseClass):

    def test_ReadAndWriteExcelFile(self):
        book = openpyxl.load_workbook('C:\\Users\\gonza\\PycharmProjects\\PracticeSofTesting\\Data\\example.xlsx')
        sheet = book.active
        Dict = {}
        #este comando de abajo es para leer una fila y columna en el excel
        cell = sheet.cell(row=1, column=1)
        #print(cell.value)
        #ahora para escribir en el excel es el siguiente comando
        #en la siguiente celda, estamos escribiendo el nombre gonzalo
        firstName = sheet.cell(row=2, column=2).value = "Gonzalo"
        firstLastName = sheet.cell(row=2, column=3).value = "Molina"
        secondName = sheet.cell(row=3, column=2).value = "Sergio"
        secondLastName = sheet.cell(row=3, column=3).value = "Martinez"
        #print(firstName)
        #el siguiente comando me permite saber cuants filas estoy utilizando del sheet1
        #print(sheet.max_row)
        #print(sheet['B1'].value)
        #con el siguiente for voy a recorrer, el archivo excel e imprimir los valores
        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == "Test Case 1":
                for j in range(2, sheet.max_column+1):
                    #print(sheet.cell(row=i, column=j).value)
                    #con el comando siguiente estoy creando el formato de diccionario que es First Name: Gonzalo,
                    #Last Name: Molina. {'First Name': 'Gonzalo', 'Last Name': 'Molina'}
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print(Dict)