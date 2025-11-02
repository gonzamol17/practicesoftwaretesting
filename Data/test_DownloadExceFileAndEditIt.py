import time

import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec




class TestDownloadExcelFileAndEditIt():

    @staticmethod
    def updateTheExcelFile(filePath, searchTerm, colName, newValue):
        file_path = "C:/Users/gonza/Downloads/download.xlsx"
        book = openpyxl.load_workbook(filePath)
        sheet = book.active
        Dict = {}

        for i in range(1,sheet.max_row+1):
            if sheet.cell(row=1, column=i).value == colName:
                Dict["col"] = i

        for i in range(1,sheet.max_row+1):
            for j in range(1,sheet.max_column+1):
                if sheet.cell(row=i, column=j).value == searchTerm:
                    Dict["row"] = i

        sheet.cell(row=Dict["row"], column=Dict["col"]).value = newValue
        book.save(file_path)


    def testDownloadExcelFileAndEditIt(self):
        # Edit the excel with updated value

        driver = webdriver.Chrome()
        url = "https://rahulshettyacademy.com/upload-download-test/index.html"
        driver.get(url)

        filePath = "C:/Users/gonza/Downloads/download.xlsx"
        fruit_name = 'Apple'
        newValue = "665"
        time.sleep(4)
        driver.find_element(By.ID, "downloadButton").click()

        #Edit the excel with updated value
        TestDownloadExcelFileAndEditIt.updateTheExcelFile(filePath, fruit_name, "price", newValue)

        time.sleep(1)
        #upload process
        file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
        file_input.send_keys(filePath)

        time.sleep(3)
        wait = WebDriverWait(driver, 5)
        toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body > div:nth-child(2)")
        wait.until(ec.visibility_of_element_located(toast_locator))
        print(driver.find_element(*toast_locator).text)

        priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
        actualPrice = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
        print("This is the actual Price: "+actualPrice)
        print("This is the new Price: "+newValue)
        assert actualPrice == newValue

        time.sleep(5)